# ============================================================
# CAMADA 2 — Módulos Funcionais
# excel/reader.py — Leitura e parsing da planilha Excel
# ============================================================

import logging
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _detect_photo_columns(rows):
    """Detecta colunas de chave e URL na aba de fotos por cabeçalho ou conteúdo."""
    import re
    key_re = re.compile(r"c[óo]digo|chave|foto do bem|key|\bid\b", re.I)
    url_re = re.compile(r"link|download|url|http", re.I)
    for r in range(min(5, len(rows))):
        row = rows[r] or []
        key_col = url_col = -1
        for c, cell in enumerate(row):
            if cell is None:
                continue
            txt = str(cell).strip()
            if not txt:
                continue
            if key_col == -1 and key_re.search(txt):
                key_col = c
            if url_re.search(txt):
                url_col = c
        if key_col != -1 and url_col != -1 and key_col != url_col:
            return key_col, url_col

    key_pat = re.compile(r"^[\w-]+\.\d+(\.\d+)?$")
    key_scores, url_scores = {}, {}
    for r in range(min(60, len(rows))):
        row = rows[r] or []
        for c, cell in enumerate(row):
            if cell is None:
                continue
            txt = str(cell).strip()
            if not txt:
                continue
            if re.match(r"^https?://", txt, re.I):
                url_scores[c] = url_scores.get(c, 0) + 1
            elif key_pat.match(txt):
                key_scores[c] = key_scores.get(c, 0) + 1
    if key_scores and url_scores:
        best_key = max(key_scores, key=key_scores.get)
        best_url = max(url_scores, key=url_scores.get)
        return best_key, best_url
    return None


def build_photo_lookup(filepath):
    """
    Constrói lookup de fotos a partir da aba 'Foto do Bem'.
    Auto-detecta colunas (ex: A=Código, C=Link). Fallback A→C, A→B, BB→DD, B→D.
    """
    lookup = {}
    try:
        from openpyxl.utils import column_index_from_string
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_name = next((s for s in wb.sheetnames if "foto" in s.lower() and "bem" in s.lower()), None)
        if not sheet_name:
            logger.warning("[CAMADA 2][excel][reader.build_photo_lookup] Aba 'Foto do Bem' não encontrada")
            wb.close()
            return lookup

        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        candidate_pairs = []
        detected = _detect_photo_columns(rows)
        if detected:
            candidate_pairs.append(detected)
        candidate_pairs.extend([
            (0, 2), (0, 1),
            (column_index_from_string("BB") - 1, column_index_from_string("DD") - 1),
            (column_index_from_string("B") - 1, column_index_from_string("D") - 1),
        ])

        import re
        for key_idx, url_idx in candidate_pairs:
            if key_idx is None or url_idx is None or key_idx == url_idx:
                continue
            found = 0
            for row in rows:
                if len(row) <= max(key_idx, url_idx):
                    continue
                key, url = row[key_idx], row[url_idx]
                if key is None or url is None:
                    continue
                key_str, url_str = str(key).strip(), str(url).strip()
                if not key_str or not url_str:
                    continue
                if not re.match(r"^https?://", url_str, re.I):
                    continue
                if re.match(r"^(c[óo]digo|chave|key|id|link|download|url|nome)", key_str, re.I):
                    continue
                if key_str not in lookup:
                    lookup[key_str] = url_str
                    found += 1
            if found > 0:
                logger.info("[CAMADA 2][excel][reader.build_photo_lookup] %d fotos (cols %s→%s)", found, key_idx, url_idx)
                break

        wb.close()
    except Exception as e:
        logger.error("[CAMADA 2][excel][reader.build_photo_lookup] %s", str(e))
    return lookup


def read_headers(filepath, sheet_index=0):
    """
    Lê os headers da planilha e retorna lista de colunas disponíveis.
    Detecta automaticamente a linha de headers (primeira linha não-vazia com texto).
    
    Returns:
        dict com:
        - headers: lista de {letter, index, name} para cada coluna com conteúdo
        - sheet_names: lista de nomes das abas
        - total_rows: total de linhas com dados
        - header_row: número da linha usada como header
        - best_sheet_idx: índice da aba detectada como principal
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        
        # Auto-detect best sheet based on data size (excluding auxiliary/config sheets)
        best_sheet_idx = 0
        max_cells = -1
        for i, ws_iter in enumerate(wb.worksheets):
            if ws_iter.title.strip().lower() in ["foto do bem", "historico", "histórico", "config", "configurações", "configuracoes"]:
                continue
            cells = ws_iter.max_row * ws_iter.max_column
            if cells > max_cells:
                max_cells = cells
                best_sheet_idx = i
                
        ws = wb.worksheets[best_sheet_idx]

        # Detectar linha de headers (procura nas primeiras 20 linhas)
        header_row = 1
        headers = []

        for row_idx in range(1, 21):
            row_headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    row_headers.append({
                        "letter": get_column_letter(col_idx),
                        "index": col_idx,
                        "name": str(cell_value).strip()
                    })
            # Usar a primeira linha que tenha pelo menos 2 colunas preenchidas
            if len(row_headers) >= 2:
                headers = row_headers
                header_row = row_idx
                break

        # Contar linhas com dados (após o header)
        total_rows = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            has_data = False
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    has_data = True
                    break
            if has_data:
                total_rows += 1

        wb.close()

        logger.info(
            "[CAMADA 2][excel][reader.read_headers] "
            "Arquivo: %s, Headers: %d colunas, Dados: %d linhas, Aba Selecionada: %s",
            filepath, len(headers), total_rows, ws.title
        )

        return {
            "status": "ok",
            "headers": headers,
            "sheet_names": sheet_names,
            "total_rows": total_rows,
            "header_row": header_row,
            "best_sheet_idx": best_sheet_idx
        }

    except Exception as e:
        logger.error("[CAMADA 2][excel][reader.read_headers] %s", str(e))
        return {"status": "error", "message": str(e)}


def read_data(filepath, sheet_index=0, header_row=1, max_rows=None, column_mappings=None):
    """
    Lê os dados da planilha a partir da linha de header. Resolvendo links de fotos da aba Foto do Bem.
    
    Returns:
        dict com:
        - rows: lista de dicts {coluna_letter: valor}
        - headers: dict {letter: name}
        - total: contagem total de linhas
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Use provided sheet_index if it's within range, otherwise auto-detect
        if sheet_index is not None and 0 <= sheet_index < len(wb.worksheets):
            ws = wb.worksheets[sheet_index]
            best_sheet_idx = sheet_index
        else:
            best_sheet_idx = 0
            max_cells = -1
            for i, ws_iter in enumerate(wb.worksheets):
                if ws_iter.title.strip().lower() in ["foto do bem", "historico", "histórico", "config", "configurações", "configuracoes"]:
                    continue
                cells = ws_iter.max_row * ws_iter.max_column
                if cells > max_cells:
                    max_cells = cells
                    best_sheet_idx = i
            ws = wb.worksheets[best_sheet_idx]

        # Ler headers
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                letter = get_column_letter(col_idx)
                headers[letter] = str(val).strip()

        # Construir lookup de fotos a partir da aba 'Foto do Bem'
        photo_lookup = build_photo_lookup(filepath)

        # Ler dados
        rows = []
        end_row = ws.max_row + 1
        if max_rows:
            end_row = min(header_row + 1 + max_rows, ws.max_row + 1)

        for row_idx in range(header_row + 1, end_row):
            row_data = {"_row_index": row_idx}
            has_data = False
            
            # Primeiro pass: ler valores brutos
            raw_row = {}
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                raw_row[letter] = val
                
            if not has_data:
                continue

            # Segundo pass: resolver fotos usando coluna A como ID do bem (chave foto)
            asset_id = raw_row.get("A")
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                val = raw_row[letter]

                is_photo_col = False
                photo_idx = None

                # 1. Verificar se está mapeado
                if column_mappings:
                    for field_name, map_info in column_mappings.items():
                        col_let = map_info.get("letter") if isinstance(map_info, dict) else map_info
                        if col_let == letter:
                            if field_name == "photo_original":
                                is_photo_col = True
                                photo_idx = "0"
                            elif field_name == "photo_spec":
                                is_photo_col = True
                                photo_idx = "1"
                            elif field_name == "photo_tag":
                                is_photo_col = True
                                photo_idx = "2"

                if not is_photo_col and (val == "Foto" or (val and str(val).strip().lower() == "foto")):
                    header_name = headers.get(letter, "").lower()
                    if any(kw in header_name for kw in ["foto do bem 2", "foto especificações", "foto especificação", "foto especificacoes", "foto especificacao"]):
                        is_photo_col = True
                        photo_idx = "1"
                    elif any(kw in header_name for kw in ["foto da tag", "foto tag", "foto da plaqueta", "foto do bem 3"]):
                        is_photo_col = True
                        photo_idx = "2"
                    elif any(kw in header_name for kw in ["foto do bem 1", "foto do ativo", "foto original", "foto do bem"]):
                        is_photo_col = True
                        photo_idx = "0"
                    else:
                        is_photo_col = True
                        photo_idx = "0"

                if is_photo_col and asset_id is not None:
                    if isinstance(asset_id, float) and asset_id == int(asset_id):
                        asset_id_str = str(int(asset_id))
                    else:
                        asset_id_str = str(asset_id).strip()

                    key = f"{asset_id_str}.{photo_idx}"
                    if key in photo_lookup:
                        val = photo_lookup[key]
                    elif val == "Foto" or (val and str(val).strip().lower() == "foto"):
                        val = None

                row_data[letter] = val
            rows.append(row_data)

        wb.close()

        logger.info(
            "[CAMADA 2][excel][reader.read_data] Lidas %d linhas de dados, resolvidas fotos", len(rows)
        )

        return {
            "status": "ok",
            "rows": rows,
            "headers": headers,
            "total": len(rows)
        }

    except Exception as e:
        logger.error("[CAMADA 2][excel][reader.read_data] %s", str(e))
        return {"status": "error", "message": str(e)}


def get_preview(filepath, sheet_index=0, num_rows=5):
    """Retorna preview dos primeiros N rows para validação visual."""
    try:
        result = read_headers(filepath, sheet_index)
        if result["status"] != "ok":
            return result

        data = read_data(filepath, result["best_sheet_idx"], result["header_row"], max_rows=num_rows)
        if data["status"] != "ok":
            return data

        return {
            "status": "ok",
            "headers": result["headers"],
            "preview_rows": data["rows"],
            "total_rows": result["total_rows"],
            "sheet_names": result["sheet_names"],
            "header_row": result["header_row"],
            "best_sheet_idx": result["best_sheet_idx"]
        }

    except Exception as e:
        logger.error("[CAMADA 2][excel][reader.get_preview] %s", str(e))
        return {"status": "error", "message": str(e)}
