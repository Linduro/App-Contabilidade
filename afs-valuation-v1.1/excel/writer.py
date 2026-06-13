# ============================================================
# CAMADA 2 — Módulos Funcionais
# excel/writer.py — Gravação de resultados nas colunas do Excel
# ============================================================

import logging
import openpyxl

logger = logging.getLogger(__name__)


def write_cell(filepath, sheet_index, row, column, value):
    """
    Grava um valor em uma célula específica da planilha.
    
    Args:
        filepath: caminho do arquivo .xlsx
        sheet_index: índice da aba (0-based)
        row: número da linha (1-based)
        column: letra ou índice da coluna
        value: valor a gravar
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.worksheets[sheet_index]
        ws.cell(row=row, column=column if isinstance(column, int) else
                openpyxl.utils.column_index_from_string(column),
                value=value)
        wb.save(filepath)
        wb.close()
        logger.info(
            "[CAMADA 2][excel][writer.write_cell] "
            "Gravado valor na linha %d, coluna %s", row, column
        )
        return {"status": "ok"}
    except Exception as e:
        logger.error("[CAMADA 2][excel][writer.write_cell] %s", str(e))
        return {"status": "error", "message": str(e)}


def write_row_results(filepath, sheet_index, row, results, column_mappings):
    """
    Grava os resultados de uma avaliação na linha correspondente.
    
    Args:
        filepath: caminho do arquivo .xlsx
        sheet_index: índice da aba
        row: número da linha
        results: dict {field_name: value}
        column_mappings: dict {field_name: {letter, index}}
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.worksheets[sheet_index]

        written = 0
        for field_name, value in results.items():
            if field_name in column_mappings and value is not None:
                map_info = column_mappings[field_name]
                if isinstance(map_info, dict):
                    col_idx = map_info["index"]
                else:
                    from openpyxl.utils import column_index_from_string
                    col_idx = column_index_from_string(str(map_info))
                ws.cell(row=row, column=col_idx, value=value)
                written += 1

        wb.save(filepath)
        wb.close()
        logger.info(
            "[CAMADA 2][excel][writer.write_row_results] "
            "Gravados %d campos na linha %d", written, row
        )
        return {"status": "ok", "fields_written": written}
    except Exception as e:
        logger.error("[CAMADA 2][excel][writer.write_row_results] %s", str(e))
        return {"status": "error", "message": str(e)}
