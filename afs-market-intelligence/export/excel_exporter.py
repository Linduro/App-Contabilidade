"""Export Excel consolidado — pronto para abordagem manual."""

import logging
import os
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)
EXPORT_DIR = Path(os.environ.get("EXPORT_DIR", "data/exports"))


class ExcelExporter:
    """Gera planilha com abas: Leads Prontos, Dead Zone, Transição, Parceiros, Metadados."""

    def __init__(self, conn):
        self.conn = conn

    def exportar(self, perfil: str = "patrimonial") -> dict:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {"status": "error", "message": "openpyxl não instalado"}

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = EXPORT_DIR / f"afs_leads_{perfil}_{ts}.xlsx"

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")

        self._aba_leads_prontos(wb, perfil, header_font, header_fill)
        self._aba_dead_zone(wb, perfil, header_font, header_fill)
        self._aba_transicao(wb, header_font, header_fill)
        self._aba_parceiros(wb, header_font, header_fill)
        self._aba_metadados(wb, perfil, header_font, header_fill)

        wb.save(filepath)
        logger.info("[export] Arquivo gerado: %s", filepath)
        return {"status": "ok", "filepath": str(filepath), "filename": filepath.name}

    def exportar_prospectos_rf(
        self,
        uf: str | None = None,
        cluster: str | None = None,
        q: str | None = None,
        cnae: str | None = None,
        porte: str | None = None,
        municipio: str | None = None,
        capital_min: float | None = None,
        capital_max: float | None = None,
        limite: int = 100_000,
    ) -> dict:
        """Exporta base prospectos_rf (ingestão Lucro Real) para Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {"status": "error", "message": "openpyxl não instalado"}

        from layers.categorization.prospect_filters import parse_filters, sql_where, load_defaults

        defaults = load_defaults().get("icp_ativo", {})
        flt = parse_filters({
            "uf": uf, "cluster": cluster, "q": q, "cnae": cnae, "porte": porte,
            "municipio": municipio,
            "capital_min": capital_min if capital_min is not None else defaults.get("capital_min"),
            "capital_max": capital_max if capital_max is not None else defaults.get("capital_max"),
        })
        where, params = sql_where(flt)
        limite = min(max(int(limite), 1), 250_000)
        sql = f"""
            SELECT cnpj_basico, cnpj_matriz, razao_social, cluster_estrategico,
                   capital_social, cnae_principal, cnae_principal_descricao,
                   uf, municipio_nome, porte, email_matriz, telefone_matriz, endereco_matriz,
                   socios_chave, qtd_estabelecimentos, score_prioridade, status_funil
            FROM prospectos_rf WHERE {where}
            ORDER BY score_prioridade DESC LIMIT ?
        """
        params.append(limite)

        try:
            rows = self.conn.execute(sql, params).fetchall()
        except Exception as e:
            return {"status": "error", "message": f"Tabela prospectos_rf indisponível: {e}"}

        if not rows:
            return {"status": "error", "message": "Nenhum prospecto com estes filtros — ajuste capital/CNAE ou rode a ingestão."}

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{uf}" if uf else ""
        filepath = EXPORT_DIR / f"afs_prospectos_lr{suffix}_{ts}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Prospectos LR"
        headers = [
            "CNPJ", "CNPJ Matriz", "Razão Social", "Cluster", "Capital Social",
            "CNAE", "Descrição CNAE", "UF", "Município", "Porte (func. proxy)",
            "E-mail", "Telefone", "Endereço", "Sócios", "Qtd Filiais", "Score", "Status Funil",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

        meta = wb.create_sheet("Metadados")
        meta_rows = [
            ("Plataforma", "AFS Market Intelligence"),
            ("Aba", "Prospectos Lucro Real (RF)"),
            ("Registros exportados", len(rows)),
            ("Capital mín.", flt.get("capital_min")),
            ("Capital máx.", flt.get("capital_max")),
            ("CNAE", cnae or ""),
            ("Porte", porte or ""),
            ("UF", uf or "Todas"),
            ("Exportado em", datetime.now().isoformat()),
        ]
        for i, (k, v) in enumerate(meta_rows, 1):
            meta.cell(row=i, column=1, value=k).font = header_font
            meta.cell(row=i, column=2, value=v)

        wb.save(filepath)
        logger.info("[export] Prospectos LR: %s (%d linhas)", filepath, len(rows))
        return {
            "status": "ok",
            "filepath": str(filepath),
            "filename": filepath.name,
            "total_exportado": len(rows),
            "filtros": flt,
        }

    def _style_header(self, ws, headers, font, fill):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    def _aba_leads_prontos(self, wb, perfil, font, fill):
        ws = wb.active
        ws.title = "Leads Prontos"
        headers = ["CNPJ", "Razão Social", "Cluster", "Score", "Decisor", "Cargo",
                   "E-mail", "Status E-mail", "LinkedIn", "Transição Regime"]
        self._style_header(ws, headers, font, fill)

        rows = self.conn.execute(
            """SELECT l.cnpj_basico, l.razao_social, l.cluster_estrategico, l.score_prioridade,
                      d.nome, d.cargo, d.email, e.status, d.linkedin_url, l.transicao_regime
               FROM leads_icp l
               LEFT JOIN decisores d ON d.lead_id = l.id
               LEFT JOIN emails_validados e ON e.decisor_id = d.id
               WHERE l.perfil_uso = ? AND (e.status IN ('validado_alta','validado_media') OR e.status IS NULL)
               ORDER BY l.score_prioridade DESC LIMIT 5000""",
            [perfil],
        ).fetchall()

        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

    def _aba_dead_zone(self, wb, perfil, font, fill):
        ws = wb.create_sheet("Dead Zone")
        headers = ["Razão Social", "Cluster", "Motivo", "Rota Recomendada", "LinkedIn",
                   "Telefone Matriz", "Endereço", "Prioridade"]
        self._style_header(ws, headers, font, fill)

        rows = self.conn.execute(
            """SELECT l.razao_social, l.cluster_estrategico, dz.motivo, dz.rota_recomendada,
                      dz.linkedin_url, dz.telefone_matriz, dz.endereco_completo, dz.prioridade
               FROM dead_zone dz JOIN leads_icp l ON l.id = dz.lead_id
               WHERE l.perfil_uso = ? ORDER BY dz.prioridade""",
            [perfil],
        ).fetchall()
        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

    def _aba_transicao(self, wb, font, fill):
        ws = wb.create_sheet("Transição Regime")
        headers = ["CNPJ", "Razão Social", "Regime Anterior", "Regime Novo", "Cluster", "Score", "Detectado Em"]
        self._style_header(ws, headers, font, fill)

        rows = self.conn.execute(
            """SELECT rt.cnpj_basico, l.razao_social, rt.regime_anterior, rt.regime_novo,
                      l.cluster_estrategico, l.score_prioridade, rt.detectado_em
               FROM regime_transicoes rt
               LEFT JOIN leads_icp l ON l.cnpj_basico = rt.cnpj_basico
               WHERE rt.lead_quente = TRUE ORDER BY rt.detectado_em DESC"""
        ).fetchall()
        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

    def _aba_parceiros(self, wb, font, fill):
        ws = wb.create_sheet("Parceiros B2B2B")
        headers = ["Banca", "Rede", "UF", "Website", "Status Parceria", "Contato"]
        self._style_header(ws, headers, font, fill)

        rows = self.conn.execute(
            "SELECT nome, rede, uf_sede, website, status_parceria, contato_parceria FROM parceiros_auditoria ORDER BY nome"
        ).fetchall()
        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

    def _aba_metadados(self, wb, perfil, font, fill):
        ws = wb.create_sheet("Metadados")
        meta = [
            ("Plataforma", "AFS Market Intelligence"),
            ("Perfil ICP", perfil),
            ("Data Export", datetime.now().isoformat()),
            ("Versão", "1.0.0"),
            ("Abordagem", "100% manual — personalização extrema"),
        ]
        for i, (k, v) in enumerate(meta, 1):
            ws.cell(row=i, column=1, value=k).font = font
            ws.cell(row=i, column=2, value=v)
